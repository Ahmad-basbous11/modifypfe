"""Generate Excel workbooks for grading rounds."""
from __future__ import annotations

from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import Exam, Material, Mark, Question


def _ensure_export_dir() -> Path:
    base = Path(settings.MEDIA_ROOT) / "exports"
    base.mkdir(parents=True, exist_ok=True)
    return base


def total_on_100(material: Material, exam: Exam, corrector_role: str) -> float:
    questions = Question.objects.filter(
        material=material, session_type=exam.session_type
    )
    max_possible = sum(q.part_mark for q in questions)
    if max_possible <= 0:
        return 0.0
    marks = Mark.objects.filter(
        exam=exam, corrector_role=corrector_role, question__in=questions
    )
    earned = sum(m.mark for m in marks)
    return round((earned / max_possible) * 100, 2)


def build_rubric_and_marks_workbook(
    *,
    material: Material,
    material_label: str,
    year_label: str,
    exams: list[Exam],
    questions: list[Question],
    corrector_role: str,
    role_label: str,
) -> str:
    """Sheet 1: rubric + one row per student with marks per part."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rubric & marks"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(
        [
            "Material",
            material_label,
            "Year",
            year_label,
            "Corrector role",
            role_label,
        ]
    )
    ws.append([])

    ws.append(["Question", "Part", "Max mark"])
    for row in ws.iter_rows(
        min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=3
    ):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    for q in questions:
        ws.append([q.question_title, q.part_title, q.part_mark])

    ws.append([])
    head = ["Exam #"] + [f"{q.question_title} / {q.part_title}" for q in questions]
    ws.append(head)
    for row in ws.iter_rows(
        min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(head)
    ):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    q_ids = [q.id for q in questions]
    mark_map = {}
    if questions:
        for m in Mark.objects.filter(
            exam__in=exams,
            corrector_role=corrector_role,
            question_id__in=q_ids,
        ):
            mark_map[(m.exam_id, m.question_id)] = m.mark

    for ex in exams:
        row = [ex.exam_number]
        for q in questions:
            row.append(mark_map.get((ex.id, q.id), ""))
        ws.append(row)

    out_dir = _ensure_export_dir()
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in material_label)[
        :80
    ]
    fname = f"rubric_marks_{safe}_{year_label}_{role_label}.xlsx".replace(
        " ", "_"
    )
    path = out_dir / fname
    wb.save(path)
    return str(path)


def build_totals_workbook(
    *,
    material: Material,
    material_label: str,
    year_label: str,
    exams: list[Exam],
    corrector_role: str,
    role_label: str,
) -> str:
    """Totals on /100 for each exam number."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Totals on 100"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["Material", material_label, "Year", year_label, "Role", role_label])
    ws.append([])
    ws.append(["Exam #", "Total / 100"])
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=2):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    for ex in exams:
        t = total_on_100(material, ex, corrector_role)
        ws.append([ex.exam_number, t])

    out_dir = _ensure_export_dir()
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in material_label)[
        :80
    ]
    fname = f"totals_100_{safe}_{year_label}_{role_label}.xlsx".replace(" ", "_")
    path = out_dir / fname
    wb.save(path)
    return str(path)


def build_final_workbook(
    *,
    material_label: str,
    year_label: str,
    exams: list[Exam],
    results: dict,
) -> str:
    """Final averages; highlight row if difference >= 10."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Final results"

    red_fill = PatternFill("solid", fgColor="FFCDD2")
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["Material", material_label, "Year", year_label])
    ws.append([])
    ws.append(["Exam #", "First /100", "Second /100", "Average", "Difference"])
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=5):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    for ex in exams:
        fr = results.get(ex.id)
        if not fr:
            continue
        diff = fr.difference
        row_idx = ws.max_row + 1
        ws.append(
            [
                ex.exam_number,
                round(fr.first_total, 2),
                round(fr.second_total, 2),
                round(fr.average, 2),
                round(diff, 2),
            ]
        )
        if diff >= 10:
            for c in ws[row_idx]:
                c.fill = red_fill

    out_dir = _ensure_export_dir()
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in material_label)[
        :80
    ]
    fname = f"final_{safe}_{year_label}.xlsx".replace(" ", "_")
    path = out_dir / fname
    wb.save(path)
    return str(path)


def build_both_correctors_totals_workbook(
    *,
    material: Material,
    material_label: str,
    year_label: str,
    exams: list[Exam],
) -> str:
    """Per exam: /100 total for first corrector and for second corrector."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Final marks both"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(
        ["Material", material_label, "Year", year_label, "Unit", "/100 per corrector"]
    )
    ws.append([])
    ws.append(["Exam #", "First corrector /100", "Second corrector /100"])
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=3):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    for ex in exams:
        t1 = total_on_100(material, ex, "first")
        t2 = total_on_100(material, ex, "second")
        ws.append([ex.exam_number, t1, t2])

    out_dir = _ensure_export_dir()
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in material_label)[
        :80
    ]
    fname = f"final_marks_both_{safe}_{year_label}.xlsx".replace(" ", "_")
    path = out_dir / fname
    wb.save(path)
    return str(path)


def build_average_only_workbook(
    *,
    material_label: str,
    year_label: str,
    exams: list[Exam],
    results: dict,
) -> str:
    """Per exam: agreed average /100 between the two correctors (after second round)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Averages"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["Material", material_label, "Year", year_label])
    ws.append([])
    ws.append(["Exam #", "Average /100 (between correctors)"])
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=2):
        for c in row:
            c.fill = header_fill
            c.font = header_font

    for ex in exams:
        fr = results.get(ex.id)
        avg = round(fr.average, 2) if fr else ""
        ws.append([ex.exam_number, avg])

    out_dir = _ensure_export_dir()
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in material_label)[
        :80
    ]
    fname = f"final_averages_{safe}_{year_label}.xlsx".replace(" ", "_")
    path = out_dir / fname
    wb.save(path)
    return str(path)
